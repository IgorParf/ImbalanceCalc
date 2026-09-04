"""Читання місячного файлу ГП у нормалізовану погодинну таблицю."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import IO, Any

import pandas as pd
from openpyxl import load_workbook

from ..exceptions import FileFormatError, ValidationError
from .schema import GROUP_DELTA_PARTS, HOURS_PER_DAY, REQUIRED_SHEETS, SHEET_COLUMNS


def _normalize(name: str) -> str:
    """Звести назву аркуша до порівнюваного вигляду."""
    name = name.replace(" ", " ").strip().rstrip(".").strip()
    return re.sub(r"\s+", " ", name).casefold()


def _match_sheet(sheet_name: str) -> str | None:
    """Знайти колонку, якій відповідає аркуш.

    Спочатку шукається точний збіг, потім — єдиний канонічний варіант, що
    починається з (обрізаної) назви аркуша: Excel скорочує довгі назви до
    31 символу, тому «Сальдовий обсяг небалансу бе...» треба зіставити з
    «Сальдовий обсяг небалансу без дельт», а не з «Сальдовий обсяг небалансу».
    """
    norm = _normalize(sheet_name)
    canonical = {_normalize(k): v for k, v in SHEET_COLUMNS.items()}

    if norm in canonical:
        return canonical[norm]

    prefix_hits = [col for key, col in canonical.items() if key.startswith(norm)]
    if len(prefix_hits) == 1:
        return prefix_hits[0]
    return None


def _parse_day(value: Any) -> date:
    """Розібрати значення колонки «Доба» (рядок ДД.ММ.РРРР або дата)."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValidationError(f"Не вдалося розібрати дату доби: {value!r}")


def _read_sheet(ws) -> tuple[dict[tuple[date, int], float], dict[date, str]]:
    """Перетворити аркуш-матрицю (доба × 24 години) на плоскі словники."""
    values: dict[tuple[date, int], float] = {}
    zones: dict[date, str] = {}
    for row in range(2, ws.max_row + 1):
        raw_day = ws.cell(row, 1).value
        if raw_day in (None, ""):
            continue
        day = _parse_day(raw_day)
        zones.setdefault(day, str(ws.cell(row, 2).value or ""))
        for hour in range(1, HOURS_PER_DAY + 1):
            cell = ws.cell(row, hour + 2).value
            values[(day, hour)] = 0.0 if cell in (None, "") else float(cell)
    return values, zones


def load_monthly_file(source: str | Path | IO[bytes]) -> pd.DataFrame:
    """Прочитати файл ГП і повернути погодинну таблицю.

    Колонки результату: ``date``, ``hour``, ``zone``, ``timestamp`` та всі
    величини з :data:`~imbalance_calc.dataio.schema.REQUIRED_COLUMNS`.
    Один рядок — одна розрахункова година.

    Відсутні аркуші команд ОСП дають нульові дельти, а окремий аркуш ``ΔΣS``
    додається до ``ΔΣW`` — див. docs/gp_file_format_variations.md.
    """
    try:
        wb = load_workbook(source, data_only=True)
    except Exception as exc:  # noqa: BLE001 — будь-яка помилка openpyxl
        raise FileFormatError(f"Не вдалося відкрити файл: {exc}") from exc

    try:
        columns: dict[str, dict[tuple[date, int], float]] = {}
        zones: dict[date, str] = {}
        for ws in wb.worksheets:
            column = _match_sheet(ws.title)
            if column is None or column in columns:
                continue
            values, sheet_zones = _read_sheet(ws)
            columns[column] = values
            for day, zone in sheet_zones.items():
                zones.setdefault(day, zone)
    finally:
        wb.close()

    missing = [name for name, col in REQUIRED_SHEETS.items() if col not in columns]
    if missing:
        raise ValidationError("У файлі бракує аркушів: " + ", ".join(missing))

    keys = sorted(set().union(*(set(v) for v in columns.values())))
    if not keys:
        raise ValidationError("Файл не містить жодної доби з даними")

    frame = pd.DataFrame(
        {
            "date": [k[0] for k in keys],
            "hour": [k[1] for k in keys],
            "zone": [zones.get(k[0], "") for k in keys],
        }
    )
    for column, series in columns.items():
        frame[column] = [series.get(key, float("nan")) for key in keys]

    # Аркушів команд ОСП може не бути — тоді дельта за місяць справді нульова.
    for column in SHEET_COLUMNS.values():
        if column not in frame.columns:
            frame[column] = 0.0

    # ΔΣS існує окремим аркушем не завжди; коли існує, дельта групи — сума
    # обох аркушів. Пропустити його означає взяти неповну дельту й отримати
    # інший рахунок без жодної помилки, тож згортаємо одразу при читанні.
    base, extra = GROUP_DELTA_PARTS
    frame[base] = frame[base] + frame.pop(extra)

    frame["timestamp"] = pd.to_datetime(frame["date"]) + pd.to_timedelta(frame["hour"] - 1, "h")
    return frame


def guess_period_label(source_name: str) -> str:
    """Витягти підпис періоду з імені файлу, якщо він там є."""
    match = re.search(r"([А-Яа-яЇїІіЄєҐґ']+)\s+(\d{4})", source_name)
    return f"{match.group(1)} {match.group(2)}" if match else ""
